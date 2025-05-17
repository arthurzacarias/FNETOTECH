import pandas as pd
import os
import time

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from datetime import datetime
from selenium.common.exceptions import NoSuchElementException

# Montar nome do arquivo Excel
agora = datetime.now()
data_hora_formatada = agora.strftime("%d-%m %Hh%M")
arquivo_excel = f"Verificação cadastro motoristas {data_hora_formatada}.xlsx"

# Criar novo arquivo Excel com cabeçalho
wb = Workbook()
ws = wb.active
ws.title = "Status cadastro de motoristas"
ws['A1'] = "Nome"
ws['B1'] = "Status"
wb.save(arquivo_excel)

# Configurações do Chrome

# Caminho para o diretório de perfil do Chrome
userdir = 'c:\\chromewithlogin'

# Configurar opções do Chrome
options = Options()
options.add_argument(f"--user-data-dir={userdir}")

# Cria o driver
driver = webdriver.Chrome(options=options)

# URL do site da amazon
url = 'https://logistics.amazon.com.br/account-management/delivery-associates?providerType=DA&providerStatus=ONBOARDING&searchStart=0&searchSize=100'

# Acessa a página
driver.get(url)  

# Aguarda o carregamento da página
wait = WebDriverWait(driver, 10)

# Clica na div que abre o dropdown
dropdown = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "af-select__container__label")))
dropdown.click()

# Aguarda o item "Integração" estar presente no DOM
integracao = wait.until(EC.presence_of_element_located((By.XPATH, "//li[contains(@class, 'af-option') and normalize-space()='Integração']")))

# Usa JavaScript para forçar o clique
driver.execute_script("arguments[0].click();", integracao)

# Aguarda o carregamento da tabela
pesquisar = wait.until(EC.presence_of_element_located((By.XPATH, "//button[contains(@class, 'af-button primary  af-search__search-button desktop') and normalize-space()='Pesquisar']"))) 

# Usa JavaScript para forçar o clique
driver.execute_script("arguments[0].click();", pesquisar)

# Aguarda o carregamento da pesquisa
time.sleep(3)

nomes_links = driver.find_elements(By.XPATH, "//table[contains(@class, 'af-table')]//a[@class='af-link']")
# Loop para abrir cada link em uma nova aba
for link in nomes_links:
    nome = link.text
    motorista = link.get_attribute("href")
    
    # Abrir em nova aba
    driver.execute_script("window.open(arguments[0]);", motorista)
    driver.switch_to.window(driver.window_handles[1])

    # Armazena o nome
    wait.until(EC.visibility_of_element_located((By.XPATH, "//div[contains(@class, 'af-row') and contains(@class, 'margin-left')]//h2[contains(@class, 'af-heading')]//b")))  
    elemento_nome = driver.find_element(By.XPATH, "//div[contains(@class, 'af-row') and contains(@class, 'margin-left')]//h2[contains(@class, 'af-heading')]//b")
    nome = elemento_nome.text.strip()

    # Aguarda o carregamento da div onbording
    botao_onboarding = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.af-expander__heading")))
    # Clica via JavaScript
    driver.execute_script("arguments[0].click();", botao_onboarding)

    # Verifica a impressão do crachá
    captura_classe = driver.find_element(By.CSS_SELECTOR, ".af-column.margin-left.onboarding-task")
    elemento = driver.find_element(By.XPATH, "//h3[contains(text(), 'Impressão de crachá')]")
    elemento_pai = elemento.find_element(By.XPATH, "./ancestor::div[contains(@class, 'onboarding-task-header')]")
    classes = elemento_pai.get_attribute("class")
    # If para verificações
    if "not-started" in classes:
        status = "Impressão do crachá pendente"
    elif "error" in classes:
        status = "Impressão do crachá mal-sucedida"

    # Verifica se o motorista assistiu os vídeos de integração
    captura_classe = driver.find_element(By.CSS_SELECTOR, ".af-column.margin-left.onboarding-task")
    elemento = driver.find_element(By.XPATH, "//h3[contains(text(), 'Vídeos de integração')]")
    elemento_pai = elemento.find_element(By.XPATH, "./ancestor::div[contains(@class, 'onboarding-task-header')]")
    classes = elemento_pai.get_attribute("class")
    # If para verificações
    if "not-started" in classes:
        status = "Visualização dos vídeos de integração pendente"


    try:
        # Verifica se o elemento "Exame toxicológico" existe
        elemento = driver.find_element(By.XPATH, "//h3[contains(text(), 'Exame toxicológico')]")
        
        # Se chegou aqui, o elemento existe, então segue com a lógica
        captura_classe = driver.find_element(By.CSS_SELECTOR, ".af-column.margin-left.onboarding-task")
        elemento_pai = elemento.find_element(By.XPATH, "./ancestor::div[contains(@class, 'onboarding-task-header')]")
        classes = elemento_pai.get_attribute("class")
        
        # If para verificações
        if "not-started" in classes:
            status = "Exame toxicológico pendente"
        elif "error" in classes:
            status = "Exame toxicológico reprovado"

    except NoSuchElementException:
        # Elemento não existe, então ignora e segue em frente
        pass

    # Verifica o status da verificação de antecedentes
    captura_classe = driver.find_element(By.CSS_SELECTOR, ".af-column.margin-left.onboarding-task")
    elemento = driver.find_element(By.XPATH, "//h3[contains(text(), 'Verificação de antecedentes')]")
    elemento_pai = elemento.find_element(By.XPATH, "./ancestor::div[contains(@class, 'onboarding-task-header')]")
    classes = elemento_pai.get_attribute("class")
    # If para verificações
    if "not-started" in classes:
        status = "Verificação de antecedentes pendente"
    elif "error" in classes:
        status = "Verificação de antecedentes reprovada"

    # Faz a verificação da carteira de motorista
    captura_classe = driver.find_element(By.CSS_SELECTOR, ".af-column.margin-left.onboarding-task")
    elemento = driver.find_element(By.XPATH, "//h3[contains(text(), 'Carteira de habilitação')]")
    elemento_pai = elemento.find_element(By.XPATH, "./ancestor::div[contains(@class, 'onboarding-task-header')]")
    classes = elemento_pai.get_attribute("class")
    # If para ver se está reprovada, aprovada, ou rejeitada
    if "not-started" in classes:
        status = "Anexo da CNH pendente"
    elif "error" in classes:
        status = "CNH reprovada"

    # Verifica se o motorista anexou a foto do crachá
    captura_classe = driver.find_element(By.CSS_SELECTOR, ".af-column.margin-left.onboarding-task")
    elemento = driver.find_element(By.XPATH, "//h3[contains(text(), 'Foto do crachá')]")
    elemento_pai = elemento.find_element(By.XPATH, "./ancestor::div[contains(@class, 'onboarding-task-header')]")
    classes = elemento_pai.get_attribute("class")
    # If para verificações
    if "not-started" in classes:
        status = "Anexo da Foto do crachá pendente"

    # Verifica se o aceitou o Agreement
    captura_classe = driver.find_element(By.CSS_SELECTOR, ".af-column.margin-left.onboarding-task")
    elemento = driver.find_element(By.XPATH, "//h3[contains(text(), 'Agreement')]")
    elemento_pai = elemento.find_element(By.XPATH, "./ancestor::div[contains(@class, 'onboarding-task-header')]")
    classes = elemento_pai.get_attribute("class")
    # If para verificações
    if "not-started" in classes:
        status = "Aceite do Agreement pendente"

    # Verifica se o motorista aceitou o convite
    captura_classe = driver.find_element(By.CSS_SELECTOR, ".af-column.margin-left.onboarding-task")
    elemento = driver.find_element(By.XPATH, "//h3[contains(text(), 'Aceitar convite')]")
    elemento_pai = elemento.find_element(By.XPATH, "./ancestor::div[contains(@class, 'onboarding-task-header')]")
    classes = elemento_pai.get_attribute("class")
    # If para ver se está pendente o convite
    if "in-progress" in classes:
        status = "Aceite do convite pendente"

    # Abrir arquivo Excel e escrever dados
    wb = load_workbook(arquivo_excel)
    ws = wb.active
    proxima_linha = ws.max_row + 1
    ws.cell(row=proxima_linha, column=1).value = nome
    ws.cell(row=proxima_linha, column=2).value = status
    wb.save(arquivo_excel)

    driver.close()  # Fecha a aba atual
    driver.switch_to.window(driver.window_handles[0])  # Retorna para a aba principal


